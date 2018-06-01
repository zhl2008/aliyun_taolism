#!/usr/bin/env python

from openpyxl import load_workbook
import sys
import os



def check_path(filepath):
	if os.path.exists(filepath):
		return True
	sys.exit('[!] filepath:%s not exists'%filepath)

wb = load_workbook("test5.xlsx")
print(wb.sheetnames)
sheet_lee = wb.get_sheet_by_name("0")
sheet_ins = wb.get_sheet_by_name("1")
sheet_haozi = wb.get_sheet_by_name("2")

sheets = [sheet_lee,sheet_ins,sheet_haozi]

res = ''

for sheet in sheets:
	projects = sheet['A']
	vulnerables = sheet['B']
	filepaths = sheet['C']
	functions = sheet['D']

	# the length of all columns should be the same
	assert len(projects)==len(vulnerables)
	assert len(vulnerables)==len(filepaths)
	assert len(filepaths)==len(functions)

	for i in range(len(projects)):
		if not i==0 and projects[i].value and vulnerables[i].value:
			# pre handle
			project = projects[i].value.strip()
			vulnerable = vulnerables[i].value.strip()
			if filepaths[i].value:
				filepath = filepaths[i].value.strip()
			else:
				filepath = ''

			if not functions[i].value:
				function = ''
			else:
				function = functions[i].value.strip()

			if filepath:
				check_path(filepath)

			if vulnerable=='yes':
				res += project + ':' + vulnerable + ',' + filepath + ',' + function + '\n'
			elif vulnerable=='no':
				res += project + ':' + vulnerable + ',,' + '\n'
			else:
				sys.exit('[!] wtf?')

	print res
	open('submit.txt','w').write(res)







