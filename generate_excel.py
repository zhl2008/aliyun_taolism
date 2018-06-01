#!/usr/bin/env python

from openpyxl import Workbook
import sys


wb = Workbook()

sheet_lee = wb.active
sheet_lee.title = "0"

sheet_ins = wb.create_sheet('1')


sheet_haozi = wb.create_sheet('2')

sheets = [sheet_lee,sheet_ins,sheet_haozi]

for sheet in sheets:
	sheet['A1'] = 'project'
	sheet['B1'] = 'vulnerable'
	sheet['C1'] = 'filepath'
	sheet['D1'] = 'function'

i = 1
j = 2
while i<181:

	col = 'A%d' %j
	sheet_lee[col] = 'p_%s' %(str(i).rjust(3,'0'))
	i += 1
	sheet_ins[col] = 'p_%s' %(str(i).rjust(3,'0'))
	i += 1
	sheet_haozi[col] = 'p_%s' %(str(i).rjust(3,'0'))
	i += 1
	j += 1

wb.save('test.xlsx')

